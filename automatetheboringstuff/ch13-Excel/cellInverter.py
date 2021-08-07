# Ch13 - Practice Project: Spreadsheet Cell Inverter
# Solution by SeungWon Bae

import openpyxl, sys

def quit(message):
    print(message)
    exit(0)

if len(sys.argv) == 2:
    try:
        filename = sys.argv[1]
    except:
        quit("Wrong argument type: please attach an Excel filename as command line argument")
else:
    quit("Insufficient arguments: please attach an Excel filename as command line argument")

print(f'Opening {filename}...')

try:
    wb = openpyxl.load_workbook(filename)
except:
    quit(f"Wrong filename: cannot open '{filename}'")

sheet = wb.active

wb2 = openpyxl.Workbook()
sheet2 = wb2.active

print(f'Inverting the row and column of the cells in the spreadsheet...')

for i, column in enumerate(sheet.columns):
    for j, cell in enumerate(column):
        # print(cell.value)
        sheet2.cell(row=i+1, column=j+1).value = cell.value

wb2.save(f'Inverted_{filename}')
