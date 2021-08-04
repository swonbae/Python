# Ch13 Practice Project - Multiplication Table Maker
# Solution by SeungWon Bae

import openpyxl, sys
from openpyxl.styles import Font

if len(sys.argv) > 1:
    num = int(sys.argv[1])
else:
    print("Insufficient argument: please add a number argument for table size")
    exit(0)

print(f'Creating multiplication table size of {num}...')

wb = openpyxl.Workbook()
sheet = wb.active

bold_font = Font(bold=True)

# Set the initial numbers for first row and first column
for i in range(1, num+1):
    sheet.cell(row=i+1, column=1).font = bold_font
    sheet.cell(row=1, column=i+1).font = bold_font
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=1, column=i+1).value = i

# Create the whole table
for i in range(2, num+2):
    for j in range(2, num+2):
        sheet.cell(row=i, column=j).value = sheet.cell(row=i, column=1).value * sheet.cell(row=1, column=j).value

# Save the spreadsheet
wb.save('multiplicationTable.xlsx')

print('Done')